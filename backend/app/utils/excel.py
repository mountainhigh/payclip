"""Excel 导入导出通用工具

基于 openpyxl 实现：
- export_rows: 把字典列表导出为 xlsx 字节流
- parse_import_file: 解析上传的 xlsx 文件为行字典列表
- build_template: 生成只含表头的空白模板
"""
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4A3F3F", end_color="4A3F3F", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def export_rows(rows: list, fields: list, sheet_name: str = "Sheet1") -> bytes:
    """把字典列表导出为 xlsx 字节流。

    Args:
        rows: 字典列表，每个字典是一条记录
        fields: [(excel_header, dict_key), ...] 定义列顺序与字段映射
        sheet_name: 工作表名称

    Returns:
        xlsx 文件的字节流
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 写表头
    headers = [f[0] for f in fields]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # 写数据行
    for row in rows:
        ws.append([_normalize_value(row.get(f[1])) for f in fields])

    # 列宽自适应（粗略）
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            val = row.get(fields[col_idx - 1][1])
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 4, 12), 40)

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template(fields: list, sheet_name: str = "Sheet1", sample_row: dict = None) -> bytes:
    """生成只含表头的空白模板（可选附一行示例）。

    Args:
        fields: [(excel_header, dict_key), ...]
        sheet_name: 工作表名称
        sample_row: 可选的示例数据字典，会作为第二行写入
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = [f[0] for f in fields]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    if sample_row:
        ws.append([_normalize_value(sample_row.get(f[1])) for f in fields])
    # 列宽
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(str(header)) + 4, 14)
    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_multi_sheet_workbook(sheets: list) -> bytes:
    """构建多 sheet 的 xlsx 字节流。

    Args:
        sheets: [(sheet_name, fields, rows), ...]
            - sheet_name: 工作表名称
            - fields: [(excel_header, dict_key), ...] 同 export_rows
            - rows: 字典列表

    Returns:
        xlsx 文件的字节流
    """
    wb = Workbook()
    # 移除默认创建的空 sheet
    default_ws = wb.active
    for sheet_name, fields, rows in sheets:
        ws = wb.create_sheet(title=sheet_name)
        headers = [f[0] for f in fields]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        for row in rows:
            ws.append([_normalize_value(row.get(f[1])) for f in fields])
        # 列宽自适应
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in rows:
                val = row.get(fields[col_idx - 1][1])
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 4, 12), 40)
        ws.freeze_panes = "A2"
    # 如果只有一个 sheet，删除默认空 sheet
    if default_ws is not None and len(wb.sheetnames) > 1:
        wb.remove(default_ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_import_file(file_bytes: bytes, required_headers: list) -> list:
    """解析上传的 xlsx 文件为行字典列表。

    Args:
        file_bytes: xlsx 文件字节流
        required_headers: 必填的表头列表（用于校验）

    Returns:
        [{header1: val1, header2: val2, ...}, ...] 行字典列表（跳过空行）

    Raises:
        ValueError: 表头缺失或文件格式错误
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"无法读取 Excel 文件: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 文件为空")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    # 校验必填表头
    missing = [h for h in required_headers if h not in headers]
    if missing:
        raise ValueError(f"Excel 缺少必要列: {', '.join(missing)}")
    result = []
    for row in rows[1:]:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            val = row[idx] if idx < len(row) else None
            item[header] = _clean_cell(val)
        result.append(item)
    return result


def _normalize_value(val):
    """导出时把 Python 值转为 Excel 友好值"""
    if val is None:
        return ""
    if isinstance(val, list):
        return ";".join(str(v) for v in val)
    if isinstance(val, bool):
        return "是" if val else "否"
    return val


def _clean_cell(val):
    """导入时清洗单元格值"""
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    return val
