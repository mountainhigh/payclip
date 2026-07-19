"""Excel 导入导出 + 收款补充账单分配 接口测试"""
import openpyxl
from io import BytesIO
from app.utils.excel import export_rows
from app.routers.customers import CUSTOMER_FIELDS


def _make_customer_xlsx(rows_data):
    """构造客户导入 xlsx。rows_data 中键用英文字段名（与 _customer_to_row 输出一致）"""
    return export_rows(rows_data, CUSTOMER_FIELDS, sheet_name="客户列表")


def test_export_customers(client, tenant_setup):
    """导出客户列表 → 验证 xlsx 可读取且包含已创建的客户"""
    headers = tenant_setup["tenant_headers"]
    # tenant_setup 已经创建了"测试客户"
    resp = client.get("/api/customers/export", headers=headers)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws["A1"].value == "客户名称"
    names = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert "测试客户" in names


def test_import_customers_new(client, tenant_setup):
    """导入一行新客户 → 验证 created_count=1"""
    headers = tenant_setup["tenant_headers"]
    rows = [{
        "name": "导入客户A", "status": "active", "is_new_customer": "是",
        "service_start_date": "2025-06-01", "region_tags": "广州;天河区",
        "introducer_type": "external", "introducer_name": "",
        "sales_person_name": "", "contact_phone": "13900000001",
        "contact_email": "a@b.com", "remark": "导入测试"
    }]
    xlsx_bytes = _make_customer_xlsx(rows)
    resp = client.post(
        "/api/customers/import",
        headers=headers,
        files={"file": ("customers.xlsx", xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["created_count"] == 1
    assert data["errors"] == []
    # 验证数据库中确实有该客户
    r = client.get("/api/customers?keyword=导入客户A", headers=headers)
    assert any(c["name"] == "导入客户A" for c in r.json()["data"]["items"])


def test_import_customers_update(client, tenant_setup):
    """导入同名客户 → 验证 updated_count=1，电话已更新"""
    headers = tenant_setup["tenant_headers"]
    rows = [{
        "name": "测试客户", "status": "active", "is_new_customer": "否",
        "service_start_date": "2025-01-01", "region_tags": "广州",
        "introducer_type": "external", "introducer_name": "",
        "sales_person_name": "", "contact_phone": "13900000099",  # 新电话
        "contact_email": "", "remark": "更新测试"
    }]
    xlsx_bytes = _make_customer_xlsx(rows)
    resp = client.post(
        "/api/customers/import",
        headers=headers,
        files={"file": ("customers.xlsx", xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["updated_count"] == 1
    assert data["created_count"] == 0
    # 验证电话已更新
    r = client.get("/api/customers?keyword=测试客户", headers=headers)
    cust = next(c for c in r.json()["data"]["items"] if c["name"] == "测试客户")
    assert cust["contact_phone"] == "13900000099"


def test_import_invalid_sales_name(client, tenant_setup):
    """导入时销售负责人不存在 → 验证 errors 含该行"""
    headers = tenant_setup["tenant_headers"]
    rows = [{
        "name": "导入客户B", "status": "active", "is_new_customer": "否",
        "service_start_date": "", "region_tags": "",
        "introducer_type": "external", "introducer_name": "",
        "sales_person_name": "不存在的人", "contact_phone": "",
        "contact_email": "", "remark": ""
    }]
    xlsx_bytes = _make_customer_xlsx(rows)
    resp = client.post(
        "/api/customers/import",
        headers=headers,
        files={"file": ("customers.xlsx", xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["success_count"] == 0
    assert len(data["errors"]) == 1
    assert "不存在的人" in data["errors"][0]["reason"]


def test_payment_allocation_update(client, tenant_setup):
    """创建无账单收款 → PUT allocations 补充账单 → 验证分配记录写入"""
    headers = tenant_setup["tenant_headers"]
    customer_id = tenant_setup["customer_id"]
    emp_id = tenant_setup["emp_id"]
    # 生成一笔账单
    r = client.post("/api/bills/generate", json={"year": 2025, "month": 6}, headers=headers)
    assert r.status_code == 200
    # 取该客户的未付账单
    r = client.get(f"/api/bills?company_id={customer_id}&status=unpaid", headers=headers)
    bills = r.json()["data"]["items"]
    assert len(bills) > 0
    bill_id = bills[0]["id"]
    receivable = float(bills[0]["receivable_amount"])
    # 创建一笔收款（不分配账单）
    r = client.post("/api/payments", json={
        "company_id": customer_id, "amount": receivable,
        "payment_date": "2025-06-15", "channel": "bank",
        "assigned_verifier_id": emp_id, "bill_allocations": []
    }, headers=headers)
    assert r.status_code == 200
    pid = r.json()["data"]["id"]
    # PUT allocations 补充账单
    r = client.put(f"/api/payments/{pid}/allocations", json={
        "bill_allocations": [{"bill_id": bill_id, "allocation_amount": receivable}]
    }, headers=headers)
    assert r.status_code == 200
    # 验证 PaymentBillAllocation 已写入
    r = client.get(f"/api/payments/{pid}", headers=headers)
    allocs = r.json()["data"]["bill_allocations"]
    assert len(allocs) == 1
    assert allocs[0]["bill_id"] == bill_id
    assert float(allocs[0]["allocation_amount"]) == receivable


def test_payment_allocation_only_pending(client, tenant_setup):
    """已通过的收款不能补充账单 → 返回 409"""
    headers = tenant_setup["tenant_headers"]
    customer_id = tenant_setup["customer_id"]
    emp_id = tenant_setup["emp_id"]
    # 创建收款
    r = client.post("/api/payments", json={
        "company_id": customer_id, "amount": 100,
        "payment_date": "2025-06-20", "channel": "cash",
        "assigned_verifier_id": emp_id, "bill_allocations": []
    }, headers=headers)
    pid = r.json()["data"]["id"]
    # 通过核对
    r = client.post(f"/api/payments/{pid}/verify", json={"action": "approve"}, headers=headers)
    assert r.status_code == 200
    # 尝试补充账单 → 应该 409
    r = client.put(f"/api/payments/{pid}/allocations", json={"bill_allocations": []}, headers=headers)
    assert r.status_code == 409
